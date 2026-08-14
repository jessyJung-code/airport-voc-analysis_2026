# -*- coding: utf-8 -*-
"""
보안검색 서비스 VOC 분석 스크립트
====================================
입력 : voc_raw.xlsx (원본 VOC 접수 데이터)
출력 : 1) dashboard_data.json   - HTML 대시보드용 집계 데이터
       2) 보안검색_VOC_분석결과.xlsx - 요약 시트 + 차트가 포함된 엑셀 리포트

분석 축 (8개)
  1. 등록채널   2. 요구유형   3. 고객유형   4. 서비스유형
  5. 비행편(터미널/항공사)   6. 시간대   7. 답변부서   8. 발생원인
  + 월별 추이, 요구유형×발생원인 교차분석
"""

import re
import json
import pandas as pd
from collections import Counter
from kiwipiepy import Kiwi
from sklearn.feature_extraction.text import TfidfVectorizer
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.chart import BarChart, PieChart, LineChart, Reference
from openpyxl.utils import get_column_letter

INPUT_PATH = "voc_raw.xlsx"
JSON_OUTPUT_PATH = "dashboard_data.json"
XLSX_OUTPUT_PATH = "보안검색_VOC_분석결과.xlsx"

# -------------------------------------------------------------------
# 1. 항공사 코드/명칭 정규화 테이블
#    "이용터미널/비행편" 컬럼은 자유기재 텍스트이므로
#    IATA 2-letter 코드 및 한글 항공사명을 매핑해 정규화한다.
# -------------------------------------------------------------------
IATA = {
    'KE': '대한항공', 'OZ': '아시아나항공', 'LJ': '진에어', '7C': '제주항공',
    'TW': '티웨이항공', 'BX': '에어부산', 'RS': '에어서울', 'ZE': '이스타항공',
    'YP': '에어프레미아', 'RF': '에어로케이', 'CZ': '중국남방항공', 'MU': '중국동방항공',
    'CA': '중국국제항공', 'MF': '하문항공', 'SC': '산동항공', 'CI': '중화항공',
    'CX': '캐세이퍼시픽', 'MM': '피치항공', 'NX': '에어마카오', 'TG': '타이항공',
    'VN': '베트남항공', 'VJ': '비엣젯항공', 'QR': '카타르항공', 'EK': '에미레이트항공',
    'EY': '에티하드항공', 'SQ': '싱가포르항공', 'TR': '스쿠트', 'AK': '에어아시아',
    'D7': '에어아시아엑스', 'FD': '타이에어아시아', 'PR': '필리핀항공', '5J': '세부퍼시픽',
    'JQ': '젯스타', 'QF': '콴타스항공', 'NZ': '뉴질랜드항공', 'UA': '유나이티드항공',
    'DL': '델타항공', 'AA': '아메리칸항공', 'AC': '에어캐나다', 'AF': '에어프랑스',
    'KL': 'KLM네덜란드항공', 'LH': '루프트한자', 'SU': '아에로플로트', 'TK': '터키항공',
    'ET': '에티오피아항공', 'GA': '가루다인도네시아항공', 'MH': '말레이시아항공',
    'SV': '사우디아항공', 'OM': '미얀마국영항공', 'HY': '우즈베키스탄항공',
    'LO': '폴란드항공', 'UL': '스리랑칸항공', 'IT': '타이거항공대만', 'HA': '하와이안항공',
    'JC': '일본에어커뮤터', 'OD': '말린도항공', 'XJ': '타이에어아시아엑스',
    'HB': '그레이터베이항공', 'ZG': '브이에어',
}
NAME_MAP = {
    '대한항공': '대한항공', '아시아나': '아시아나항공', '진에어': '진에어', '제주항공': '제주항공',
    '티웨이': '티웨이항공', '에어부산': '에어부산', '에어서울': '에어서울', '이스타': '이스타항공',
    '에어프레미아': '에어프레미아', '비어젯': '비엣젯항공', '비엣젯': '비엣젯항공', '베트남항공': '베트남항공',
    '중국남방': '중국남방항공', '중국동방': '중국동방항공', '중화항공': '중화항공', '산동항공': '산동항공',
    '캐세이': '캐세이퍼시픽', '타이항공': '타이항공', '카타르': '카타르항공',
    '에미레이트': '에미레이트항공', '에티하드': '에티하드항공', '에티히드': '에티하드항공',
    '싱가포르항공': '싱가포르항공', '필리핀항공': '필리핀항공', '세부항공': '세부퍼시픽',
    '유나이티드': '유나이티드항공', '델타': '델타항공', '아메리칸': '아메리칸항공',
    '에어캐나다': '에어캐나다', '에어프랑스': '에어프랑스', 'KLM': 'KLM네덜란드항공',
    '루프트한자': '루프트한자', '터키항공': '터키항공', '에티오피아': '에티오피아항공',
    '가루다': '가루다인도네시아항공', '말레이시아항공': '말레이시아항공', '피치': '피치항공',
    '에어마카오': '에어마카오', '중화': '중화항공', '에어아시아': '에어아시아',
    '스카이 앙코르': '스카이앙코르항공', '아에로멕시코': '아에로멕시코',
}


def load_and_prepare(path: str) -> pd.DataFrame:
    """원본 VOC 데이터를 읽고 분석에 필요한 파생 컬럼을 생성한다."""
    df = pd.read_excel(path)

    # 일시 컬럼 파싱
    df['등록일시_dt'] = pd.to_datetime(df['등록일시'], errors='coerce')
    df['접수일시_dt'] = pd.to_datetime(df['접수일시'], errors='coerce')
    df['완료일시_dt'] = pd.to_datetime(df['완료일시'], errors='coerce')
    df['처리시간(h)'] = (df['완료일시_dt'] - df['접수일시_dt']).dt.total_seconds() / 3600
    df['연월'] = df['등록일시_dt'].dt.to_period('M').astype(str)

    # 시간대 구간화 (3시간 단위)
    def timebin(h):
        if pd.isna(h):
            return "미상"
        h = int(h)
        edges = [0, 3, 6, 9, 12, 15, 18, 21, 24]
        for i in range(len(edges) - 1):
            if edges[i] <= h < edges[i + 1]:
                return f"{edges[i]:02d}-{edges[i+1]:02d}시"
        return "미상"
    df['시간대'] = df['등록일시_dt'].dt.hour.apply(timebin)

    # 터미널 구분
    def parse_terminal(v):
        if pd.isna(v):
            return "미기재"
        v = str(v)
        if '제1' in v:
            return "제1여객터미널"
        if '제2' in v:
            return "제2여객터미널"
        return "기타/미기재"
    df['터미널'] = df['이용터미널/비행편'].apply(parse_terminal)

    # 항공사 정규화 ("터미널 / 편명" 또는 자유기재에서 코드/명칭 추출)
    def extract_part(v):
        if pd.isna(v):
            return None
        s = str(v)
        return s.split('/')[-1].strip() if '/' in s else s.strip()

    def classify_airline(part):
        if part is None or isinstance(part, float) or part in ('-', '', '미정', '예약 예정'):
            return None
        up = part.upper().strip()
        m = re.match(r'^([A-Z]{2})[\s\-]?\d', up)
        if m and m.group(1) in IATA:
            return IATA[m.group(1)]
        for k, v2 in NAME_MAP.items():
            if k in part:
                return v2
        return None

    df['_part'] = df['이용터미널/비행편'].apply(extract_part)
    df['항공사'] = df['_part'].apply(classify_airline)

    # 서비스유형 / 답변부서: 최하위 분류명만 추출 (예: "보안/검색 > 검색대 검색" -> "검색대 검색")
    def last_segment(v):
        if pd.isna(v):
            return "미기재"
        parts = str(v).split('>')
        return parts[-1].strip() if len(parts) > 1 else str(v).strip()

    df['서비스유형_간략'] = df['서비스유형'].apply(last_segment)
    df['답변부서_간략'] = df['답변부서'].apply(last_segment)

    return df


# -------------------------------------------------------------------
# 키워드(형태소) 분석 - 제목 / 내용 텍스트에서 명사 키워드 추출
#   kiwipiepy(Kiwi) 형태소 분석기로 일반명사(NNG)/고유명사(NNP)만 추출하고,
#   길이 2자 미만 및 불용어(문맥상 의미 없는 범용 단어)는 제외한다.
# -------------------------------------------------------------------
_KIWI = Kiwi()

STOPWORDS = {
    '해당', '생각', '안녕', '가능', '필요', '관련', '경우', '정도', '사항', '이해', '당시',
    '오전', '오후', '내용', '처음', '본인', '부탁', '감사', '인천공항', '인천국제공항',
    '저희', '제가', '우리', '당일', '오늘', '어제', '내일', '자체', '정말', '너무', '자꾸',
    '계속', '일부', '전체', '모든', '아무', '상태', '이번', '다음', '마지막', '시작', '종료',
    '완료', '진행', '현재', '기타', '이상', '이하', '이후', '이전',
}


def extract_keywords(text) -> list:
    """텍스트에서 명사 키워드 리스트를 반환한다."""
    if pd.isna(text):
        return []
    tokens = _KIWI.tokenize(str(text))
    return [
        t.form for t in tokens
        if t.tag in ('NNG', 'NNP') and len(t.form) >= 2 and t.form not in STOPWORDS
    ]


def keyword_counter(series: pd.Series) -> Counter:
    cnt = Counter()
    for text in series.dropna():
        cnt.update(extract_keywords(text))
    return cnt


def build_keyword_aggregates(df: pd.DataFrame) -> dict:
    """제목/내용 키워드 빈도 및 요구유형(불편불만 vs 칭찬격려)별 비교 키워드를 집계한다."""
    out = {}

    title_cnt = keyword_counter(df['제목'])
    out['title_keywords'] = [{"name": k, "value": v} for k, v in title_cnt.most_common(20)]

    content_cnt = keyword_counter(df['내용'])
    out['content_keywords'] = [{"name": k, "value": v} for k, v in content_cnt.most_common(25)]

    complaint = df[df['요구유형'] == '불편불만']
    praise = df[df['요구유형'] == '칭찬격려']
    complaint_cnt = keyword_counter(complaint['내용'])
    praise_cnt = keyword_counter(praise['내용'])

    out['keywords_complaint'] = [{"name": k, "value": v} for k, v in complaint_cnt.most_common(15)]
    out['keywords_praise'] = [{"name": k, "value": v} for k, v in praise_cnt.most_common(15)]

    return out


# -------------------------------------------------------------------
# TF-IDF 키워드 분석
#   단순 빈도(Counter) 분석은 "직원", "보안"처럼 어디서나 자주 등장하는
#   단어가 항상 상위를 차지한다. TF-IDF는 각 문서(접수 건)별 등장 비중과
#   전체 문서군에서의 희소성을 함께 반영해 "그 카테고리에서 유독 두드러지는"
#   키워드를 잡아낸다.
# -------------------------------------------------------------------
def _tfidf_vectorizer(min_df=1):
    return TfidfVectorizer(
        tokenizer=extract_keywords,
        preprocessor=lambda x: x,  # 커스텀 토크나이저를 그대로 사용 (소문자화 등 생략)
        token_pattern=None,
        min_df=min_df,
    )


def tfidf_overall_top(series: pd.Series, top_n=20, min_df=3) -> list:
    """문서(행) 단위 TF-IDF 점수를 모두 합산해 전체 상위 키워드를 구한다.
    단순 빈도수 상위와는 다르게, 여러 문서에 걸쳐 고르게 중요한 단어가 상위에 온다."""
    docs = series.dropna().astype(str).tolist()
    if len(docs) < 5:
        return []
    vec = _tfidf_vectorizer(min_df=min_df)
    X = vec.fit_transform(docs)
    scores = X.sum(axis=0).A1
    terms = vec.get_feature_names_out()
    top = sorted(zip(terms, scores), key=lambda x: -x[1])[:top_n]
    return [{"name": t, "value": round(float(s), 2)} for t, s in top]


def tfidf_distinctive_by_group(df: pd.DataFrame, group_col: str, text_col: str,
                                groups: list, top_n=12) -> dict:
    """그룹(예: 요구유형)별로 텍스트를 하나의 문서로 합친 뒤 TF-IDF를 계산해,
    다른 그룹 대비 해당 그룹에서 유독 두드러지는(distinctive) 키워드를 추출한다."""
    class_docs = []
    valid_groups = []
    for g in groups:
        texts = df[df[group_col] == g][text_col].dropna().astype(str).tolist()
        if texts:
            class_docs.append(" ".join(texts))
            valid_groups.append(g)
    if len(class_docs) < 2:
        return {}
    vec = _tfidf_vectorizer(min_df=1)
    X = vec.fit_transform(class_docs)
    terms = vec.get_feature_names_out()
    out = {}
    for i, g in enumerate(valid_groups):
        row = X[i].toarray().flatten()
        top = sorted(zip(terms, row), key=lambda x: -x[1])[:top_n]
        out[g] = [{"name": t, "value": round(float(s), 3)} for t, s in top if s > 0]
    return out


def build_tfidf_aggregates(df: pd.DataFrame) -> dict:
    """제목/내용 TF-IDF 키워드 및 요구유형별 대표(distinctive) 키워드를 집계한다."""
    out = {
        'tfidf_title': tfidf_overall_top(df['제목'], top_n=20, min_df=2),
        'tfidf_content': tfidf_overall_top(df['내용'], top_n=25, min_df=3),
        'tfidf_by_request_type': tfidf_distinctive_by_group(
            df, group_col='요구유형', text_col='내용',
            groups=['불편불만', '상담문의', '의견제안', '칭찬격려'], top_n=10,
        ),
    }
    return out


def build_aggregates(df: pd.DataFrame) -> dict:
    """8개 분석 축 + 추가 지표를 집계하여 dict로 반환한다."""

    def vc(col, top=None, fillna="미기재"):
        s = df[col].fillna(fillna).value_counts()
        if top:
            s = s.head(top)
        return [{"name": str(k), "value": int(v)} for k, v in s.items()]

    out = {}

    out['meta'] = {
        'total': int(len(df)),
        'date_min': str(df['등록일시_dt'].min().date()),
        'date_max': str(df['등록일시_dt'].max().date()),
        'complaint_ratio': round(100 * (df['요구유형'] == '불편불만').mean(), 1),
        'avg_duration_h': round(df['처리시간(h)'].mean(), 1),
        'median_duration_h': round(df['처리시간(h)'].median(), 1),
        'praise_count': int((df['요구유형'] == '칭찬격려').sum()),
    }

    out['channel'] = vc('등록채널')
    out['request_type'] = vc('요구유형')
    out['customer_type'] = vc('고객유형', top=8)
    out['service_type'] = vc('서비스유형_간략', top=10)
    out['terminal'] = vc('터미널')

    airline = vc('항공사', top=10, fillna="__NA__")
    out['airline'] = [x for x in airline if x['name'] != "__NA__"]

    time_order = ["00-03시", "03-06시", "06-09시", "09-12시",
                  "12-15시", "15-18시", "18-21시", "21-24시"]
    tvc = df['시간대'].value_counts()
    out['timebin'] = [{"name": t, "value": int(tvc.get(t, 0))} for t in time_order]

    out['dept'] = vc('답변부서_간략', top=8)
    out['cause'] = vc('발생원인')

    mvc = df.groupby('연월').size()
    out['monthly'] = [{"month": str(k), "value": int(v)} for k, v in mvc.items()]

    ct = pd.crosstab(df['요구유형'], df['발생원인'].fillna('미기재'))
    out['cross_request_cause'] = {
        "requests": list(ct.index),
        "causes": list(ct.columns),
        "matrix": ct.values.tolist(),
    }

    out.update(build_keyword_aggregates(df))
    out.update(build_tfidf_aggregates(df))

    return out


# ---------------------------------------------------------------------------
# 엑셀 리포트 생성
# ---------------------------------------------------------------------------
NAVY = "0B1F3A"
TEAL = "0F9B8E"
LIGHT_TEAL = "DFF3F0"
HEADER_FONT = Font(name="Arial", bold=True, color="FFFFFF", size=11)
TITLE_FONT = Font(name="Arial", bold=True, color=NAVY, size=14)
BODY_FONT = Font(name="Arial", size=10.5)
HEADER_FILL = PatternFill("solid", fgColor=NAVY)
STRIPE_FILL = PatternFill("solid", fgColor=LIGHT_TEAL)
THIN = Side(style="thin", color="C9D6E8")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)


def _write_table(ws, headers, rows, start_row=1, start_col=1):
    """헤더 + 데이터 행을 스타일 적용해서 기록하고, 데이터 범위를 반환한다."""
    for j, h in enumerate(headers):
        c = ws.cell(row=start_row, column=start_col + j, value=h)
        c.font = HEADER_FONT
        c.fill = HEADER_FILL
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.border = BORDER
    for i, row in enumerate(rows, start=1):
        for j, val in enumerate(row):
            c = ws.cell(row=start_row + i, column=start_col + j, value=val)
            c.font = BODY_FONT
            c.border = BORDER
            if i % 2 == 0:
                c.fill = STRIPE_FILL
            if isinstance(val, (int, float)) and j > 0:
                c.number_format = "#,##0"
    return start_row, start_row + len(rows)


def _autofit(ws, widths):
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w


def build_excel_report(agg: dict, out_path: str):
    wb = Workbook()

    # ---------------- 요약 시트 ----------------
    ws = wb.active
    ws.title = "요약"
    ws["B2"] = "보안검색 서비스 VOC 분석 결과"
    ws["B2"].font = Font(name="Arial", bold=True, color=NAVY, size=18)
    ws["B3"] = f"분석기간: {agg['meta']['date_min']} ~ {agg['meta']['date_max']}  |  전체 접수: {agg['meta']['total']:,}건"
    ws["B3"].font = Font(name="Arial", size=11, color="44546A")

    kpis = [
        ("전체 VOC 접수", f"{agg['meta']['total']:,}건"),
        ("불편불만 비율", f"{agg['meta']['complaint_ratio']}%"),
        ("평균 처리시간", f"{agg['meta']['avg_duration_h']}시간"),
        ("처리시간 중앙값", f"{agg['meta']['median_duration_h']}시간"),
        ("칭찬·격려 건수", f"{agg['meta']['praise_count']:,}건"),
    ]
    row = 5
    for label, val in kpis:
        ws.cell(row=row, column=2, value=label).font = Font(name="Arial", bold=True, size=10.5, color=NAVY)
        ws.cell(row=row, column=3, value=val).font = Font(name="Arial", size=12, bold=True)
        row += 1

    ws.cell(row=row + 1, column=2, value="시트 구성 안내").font = Font(name="Arial", bold=True, size=11, color=NAVY)
    sheet_names = ["등록채널", "요구유형", "고객유형", "서비스유형", "터미널_항공사",
                   "시간대", "답변부서", "발생원인", "월별추이", "교차분석(요구유형x발생원인)",
                   "키워드분석(제목/내용)", "TFIDF키워드분석"]
    for i, s in enumerate(sheet_names, start=1):
        ws.cell(row=row + 1 + i, column=2, value=f"· {s}").font = BODY_FONT
    _autofit(ws, [3, 30, 20, 14])

    # ---------------- 단순 카테고리 시트 생성 헬퍼 ----------------
    def make_category_sheet(name, data_list, value_label, chart_type="bar"):
        sh = wb.create_sheet(name)
        sh["A1"] = name
        sh["A1"].font = TITLE_FONT
        headers = [name.replace("_", "/"), value_label]
        rows = [[d["name"], d["value"]] for d in data_list]
        top, bottom = _write_table(sh, headers, rows, start_row=3, start_col=1)
        _autofit(sh, [30, 14])

        n = len(rows)
        if chart_type == "bar":
            chart = BarChart()
            chart.type = "bar"
            chart.style = 10
            chart.title = f"{name} 분석"
            chart.y_axis.title = value_label
            data_ref = Reference(sh, min_col=2, min_row=3, max_row=3 + n)
            cats_ref = Reference(sh, min_col=1, min_row=4, max_row=3 + n)
            chart.add_data(data_ref, titles_from_data=True)
            chart.set_categories(cats_ref)
            chart.height, chart.width = 9, 16
        else:
            chart = PieChart()
            chart.title = f"{name} 비중"
            data_ref = Reference(sh, min_col=2, min_row=3, max_row=3 + n)
            cats_ref = Reference(sh, min_col=1, min_row=4, max_row=3 + n)
            chart.add_data(data_ref, titles_from_data=True)
            chart.set_categories(cats_ref)
            chart.height, chart.width = 9, 14
        sh.add_chart(chart, f"D3")
        return sh

    make_category_sheet("등록채널", agg["channel"], "건수", "bar")
    make_category_sheet("요구유형", agg["request_type"], "건수", "pie")
    make_category_sheet("고객유형", agg["customer_type"], "건수", "bar")
    make_category_sheet("서비스유형", agg["service_type"], "건수", "bar")

    # 터미널+항공사 결합 시트
    sh = wb.create_sheet("터미널_항공사")
    sh["A1"] = "터미널별 접수 현황"
    sh["A1"].font = TITLE_FONT
    rows_t = [[d["name"], d["value"]] for d in agg["terminal"]]
    _write_table(sh, ["터미널", "건수"], rows_t, start_row=3, start_col=1)
    chart1 = PieChart()
    chart1.title = "터미널별 비중"
    n1 = len(rows_t)
    chart1.add_data(Reference(sh, min_col=2, min_row=3, max_row=3 + n1), titles_from_data=True)
    chart1.set_categories(Reference(sh, min_col=1, min_row=4, max_row=3 + n1))
    chart1.height, chart1.width = 8, 12
    sh.add_chart(chart1, "D3")

    start_row2 = 3 + n1 + 3
    sh.cell(row=start_row2, column=1, value="항공사별 접수 TOP 10 (식별 가능 건)").font = TITLE_FONT
    rows_a = [[d["name"], d["value"]] for d in agg["airline"]]
    _write_table(sh, ["항공사", "건수"], rows_a, start_row=start_row2 + 1, start_col=1)
    chart2 = BarChart()
    chart2.title = "항공사별 VOC 접수 TOP 10"
    n2 = len(rows_a)
    chart2.add_data(Reference(sh, min_col=2, min_row=start_row2 + 1, max_row=start_row2 + 1 + n2), titles_from_data=True)
    chart2.set_categories(Reference(sh, min_col=1, min_row=start_row2 + 2, max_row=start_row2 + 1 + n2))
    chart2.height, chart2.width = 9, 16
    sh.add_chart(chart2, f"D{start_row2 + 1}")
    _autofit(sh, [26, 12])

    make_category_sheet("시간대", agg["timebin"], "건수", "bar")
    make_category_sheet("답변부서", agg["dept"], "건수", "bar")
    make_category_sheet("발생원인", agg["cause"], "건수", "bar")

    # 월별 추이 (line chart)
    sh = wb.create_sheet("월별추이")
    sh["A1"] = "월별 VOC 접수 추이"
    sh["A1"].font = TITLE_FONT
    rows_m = [[d["month"], d["value"]] for d in agg["monthly"]]
    _write_table(sh, ["연월", "건수"], rows_m, start_row=3, start_col=1)
    n = len(rows_m)
    line = LineChart()
    line.title = "월별 VOC 접수 추이"
    line.y_axis.title = "건수"
    line.add_data(Reference(sh, min_col=2, min_row=3, max_row=3 + n), titles_from_data=True)
    line.set_categories(Reference(sh, min_col=1, min_row=4, max_row=3 + n))
    line.height, line.width = 9, 22
    sh.add_chart(line, "D3")
    _autofit(sh, [12, 12])

    # 교차분석
    sh = wb.create_sheet("교차분석")
    sh["A1"] = "요구유형 x 발생원인 교차분석"
    sh["A1"].font = TITLE_FONT
    causes = agg["cross_request_cause"]["causes"]
    requests = agg["cross_request_cause"]["requests"]
    matrix = agg["cross_request_cause"]["matrix"]
    headers = ["요구유형"] + causes
    rows_x = [[requests[i]] + matrix[i] for i in range(len(requests))]
    _write_table(sh, headers, rows_x, start_row=3, start_col=1)
    n = len(rows_x)
    bar = BarChart()
    bar.type = "col"
    bar.grouping = "stacked"
    bar.overlap = 100
    bar.title = "요구유형별 발생원인 분포 (누적)"
    data_ref = Reference(sh, min_col=2, max_col=1 + len(causes), min_row=3, max_row=3 + n)
    cats_ref = Reference(sh, min_col=1, min_row=4, max_row=3 + n)
    bar.add_data(data_ref, titles_from_data=True)
    bar.set_categories(cats_ref)
    bar.height, bar.width = 10, 20
    sh.add_chart(bar, f"A{3 + n + 3}")
    _autofit(sh, [12] + [12] * len(causes))

    # ---------------- 키워드 분석 시트 ----------------
    sh = wb.create_sheet("키워드분석")
    sh["A1"] = "제목 키워드 TOP 20"
    sh["A1"].font = TITLE_FONT
    rows_tk = [[d["name"], d["value"]] for d in agg["title_keywords"]]
    _write_table(sh, ["키워드", "빈도"], rows_tk, start_row=3, start_col=1)
    n_tk = len(rows_tk)
    c1 = BarChart()
    c1.title = "제목 키워드 빈도 TOP 20"
    c1.add_data(Reference(sh, min_col=2, min_row=3, max_row=3 + n_tk), titles_from_data=True)
    c1.set_categories(Reference(sh, min_col=1, min_row=4, max_row=3 + n_tk))
    c1.height, c1.width = 10, 16
    sh.add_chart(c1, "D3")

    start2 = 3 + n_tk + 3
    sh.cell(row=start2, column=1, value="내용 키워드 TOP 25").font = TITLE_FONT
    rows_ck = [[d["name"], d["value"]] for d in agg["content_keywords"]]
    _write_table(sh, ["키워드", "빈도"], rows_ck, start_row=start2 + 1, start_col=1)
    n_ck = len(rows_ck)
    c2 = BarChart()
    c2.title = "내용 키워드 빈도 TOP 25"
    c2.add_data(Reference(sh, min_col=2, min_row=start2 + 1, max_row=start2 + 1 + n_ck), titles_from_data=True)
    c2.set_categories(Reference(sh, min_col=1, min_row=start2 + 2, max_row=start2 + 1 + n_ck))
    c2.height, c2.width = 12, 16
    sh.add_chart(c2, f"D{start2 + 1}")

    start3 = start2 + n_ck + 4
    sh.cell(row=start3, column=1,
            value="불편불만 vs 칭찬격려 - 내용 키워드 비교 (TOP 15)").font = TITLE_FONT
    comp = agg["keywords_complaint"]
    prai = agg["keywords_praise"]
    max_len = max(len(comp), len(prai))
    rows_cmp = []
    for i in range(max_len):
        c_name, c_val = (comp[i]["name"], comp[i]["value"]) if i < len(comp) else ("", "")
        p_name, p_val = (prai[i]["name"], prai[i]["value"]) if i < len(prai) else ("", "")
        rows_cmp.append([c_name, c_val, p_name, p_val])
    _write_table(sh, ["불편불만 키워드", "빈도", "칭찬격려 키워드", "빈도"], rows_cmp,
                 start_row=start3 + 1, start_col=1)
    _autofit(sh, [16, 10, 16, 10])

    # ---------------- TF-IDF 키워드 분석 시트 ----------------
    sh = wb.create_sheet("TFIDF키워드분석")
    sh["A1"] = "제목 TF-IDF 키워드 TOP 20"
    sh["A1"].font = TITLE_FONT
    sh["D1"] = "※ 단순빈도와 달리, 여러 접수 건에 걸쳐 고르게 중요한 단어가 상위에 옵니다"
    sh["D1"].font = Font(name="Arial", italic=True, size=9, color="7A7A7A")
    rows_tt = [[d["name"], d["value"]] for d in agg.get("tfidf_title", [])]
    _write_table(sh, ["키워드", "TF-IDF 점수"], rows_tt, start_row=3, start_col=1)
    n_tt = len(rows_tt)
    c1 = BarChart()
    c1.title = "제목 TF-IDF 키워드 TOP 20"
    c1.add_data(Reference(sh, min_col=2, min_row=3, max_row=3 + n_tt), titles_from_data=True)
    c1.set_categories(Reference(sh, min_col=1, min_row=4, max_row=3 + n_tt))
    c1.height, c1.width = 10, 16
    sh.add_chart(c1, "D3")

    start2 = 3 + n_tt + 3
    sh.cell(row=start2, column=1, value="내용 TF-IDF 키워드 TOP 25").font = TITLE_FONT
    rows_ct = [[d["name"], d["value"]] for d in agg.get("tfidf_content", [])]
    _write_table(sh, ["키워드", "TF-IDF 점수"], rows_ct, start_row=start2 + 1, start_col=1)
    n_ct = len(rows_ct)
    c2 = BarChart()
    c2.title = "내용 TF-IDF 키워드 TOP 25"
    c2.add_data(Reference(sh, min_col=2, min_row=start2 + 1, max_row=start2 + 1 + n_ct), titles_from_data=True)
    c2.set_categories(Reference(sh, min_col=1, min_row=start2 + 2, max_row=start2 + 1 + n_ct))
    c2.height, c2.width = 12, 16
    sh.add_chart(c2, f"D{start2 + 1}")

    start3t = start2 + n_ct + 4
    sh.cell(row=start3t, column=1,
            value="요구유형별 대표(distinctive) 키워드 TOP 10 - TF-IDF 기준").font = TITLE_FONT
    by_group = agg.get("tfidf_by_request_type", {})
    group_order = ["불편불만", "상담문의", "의견제안", "칭찬격려"]
    headers = []
    for g in group_order:
        headers += [f"{g} 키워드", "점수"]
    max_len2 = max((len(by_group.get(g, [])) for g in group_order), default=0)
    rows_g = []
    for i in range(max_len2):
        row = []
        for g in group_order:
            items = by_group.get(g, [])
            if i < len(items):
                row += [items[i]["name"], items[i]["value"]]
            else:
                row += ["", ""]
        rows_g.append(row)
    _write_table(sh, headers, rows_g, start_row=start3t + 1, start_col=1)
    _autofit(sh, [14, 8, 14, 8, 14, 8, 14, 8])

    wb.save(out_path)


if __name__ == "__main__":
    df = load_and_prepare(INPUT_PATH)
    aggregates = build_aggregates(df)

    with open(JSON_OUTPUT_PATH, "w", encoding="utf-8") as f:
        json.dump(aggregates, f, ensure_ascii=False, indent=2)
    print(f"[OK] {JSON_OUTPUT_PATH} 저장 완료")

    build_excel_report(aggregates, XLSX_OUTPUT_PATH)
    print(f"[OK] {XLSX_OUTPUT_PATH} 저장 완료")

    print("\n=== 요약 ===")
    print(f"전체 접수: {aggregates['meta']['total']:,}건")
    print(f"분석기간 : {aggregates['meta']['date_min']} ~ {aggregates['meta']['date_max']}")
    print(f"불편불만 비율: {aggregates['meta']['complaint_ratio']}%")
    print(f"평균 처리시간: {aggregates['meta']['avg_duration_h']}시간 (중앙값 {aggregates['meta']['median_duration_h']}시간)")
